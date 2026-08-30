import csv
import io
from typing import Iterable

from django.http import HttpResponse

from extractor.models import FacebookPage


def export_pages_to_csv(queryset: Iterable[FacebookPage]) -> HttpResponse:
    """Exports FacebookPage queryset to a downloadable CSV response."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # Headers
    writer.writerow(["ID", "URL", "Name", "Followers", "Status", "Updated At"])

    for page in queryset:
        writer.writerow(
            [
                page.id,
                page.url,
                page.name,
                page.followers,
                page.status,
                page.updated_at.strftime("%Y-%m-%d %H:%M:%S") if page.updated_at else "",
            ]
        )

    response = HttpResponse(
        output.getvalue().encode("utf-8-sig"),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = 'attachment; filename="facebook_fanpages.csv"'
    return response


def export_pages_to_excel(queryset: Iterable[FacebookPage]) -> HttpResponse:
    """Exports FacebookPage queryset to a formatted Excel (.xlsx) file."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        return export_pages_to_csv(queryset)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facebook Pages"

    # Header styling
    headers = ["ID", "URL", "Name", "Followers", "Status", "Last Updated"]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1877F2", end_color="1877F2", fill_type="solid"
    )  # Facebook Blue
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(
        ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)), 1
    ):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for row_idx, page in enumerate(queryset, start=2):
        ws.append(
            [
                page.id,
                page.url,
                page.name,
                page.followers,
                page.status,
                page.updated_at.strftime("%Y-%m-%d %H:%M:%S") if page.updated_at else "",
            ]
        )

        # Format followers cell as integer number
        ws.cell(row=row_idx, column=4).number_format = "#,##0"

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="facebook_fanpages.xlsx"'
    return response
